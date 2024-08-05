import mysql.connector
import openpyxl 

def get_cell_value(row, column):
    value = sheet.cell(row=row, column=column).value
    try:
        return float(value)
    except (ValueError, TypeError):
        return None

file = "MyFoodData_Nutrition_Facts_SpreadSheet_Release_1-4.xlsx"

mydb = mysql.connector.connect(
  host="localhost",
  user="root",
  password="student",
  database = "HealthNCare2_0"
)

cursor = mydb.cursor()
workbook = openpyxl.load_workbook(file)
sheet = workbook.active 

cell = sheet.cell(row = 5, column = 2)


sql = """INSERT INTO myFoodDataTable(
    ID ,
	name ,
	foodGroup ,
	calories  ,
	fat  ,
	protein  ,
	carbs  ,
	sugars  ,
	fiber  ,
	colesterol  ,
	saturatedFats  , 
	calcium  ,
	iron  ,
	potassium  ,
	magnesium  ,
	vitaminAIU  ,
	vitaminARAE  ,
	vitaminC  ,
	vitaminB12  ,
	vitaminD  ,
	vitaminEAlphaTocopheral  ,
	water  ,
	omega3s  ,
	omega6s  ,
	PRALscore  ,
	transFatAcids  ,
	solubleFiber  ,
	insolubleFiber  ,
	sucrose  ,
	glucoseDextrose  ,
	Fructose  ,
	Lactose  ,
	Maltose  ,
	Galactose  ,
	starch  ,
	phosphorus  ,
	sodium  ,
	zinc  ,
	copper  ,
	maganese  ,
	selenium  ,
	flouride  ,
	Molybdenum  ,
	chlorine  ,
	ThiaminB1  ,
	RiboflavinB2  ,
	niacinB3  ,
	PantothenicAcid  ,
	vitaminB6  ,
	biotin  ,
	folate  ,
	folicAcid  ,
	foodFolate  ,
	folateDFE  ,
	choline  ,
	beatine  ,
	retinol  ,
	caroteneBeta  ,
	caroteneAlpha  ,
	lycopene  ,
	luteinZeaxanthin  ,
	vitaminK  , 
	Dihydrophylloquinone  ,
	Menaquinone4  ,
	Tryptophan  ,
	Threonine  ,
	Isoleucine  ,
	Leucine  ,
	Lysine  ,
	Methionine  ,
	Cystine  ,
	Phenylalanine  ,
	Tyrosine  ,
	Valine  ,
	Arginine  ,
	Histidine  ,
	Alanine  ,
	AsparticAcid  ,
	GlutamicAcid  ,
	Glycine  ,
	Proline  ,
	Serine  ,
	Hydroxyproline  ,
	Alcohol  ,
	Caffeine  ,
	Theobromine  ,
	singleServingWeight  ,
	servingDesc 
 ) 
 
 VALUES(
    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
)
	
  
  """

columnsToAvoid = (22, 23, 38, 65, 66, 67, 71, 72, 73, 74, 75, 76)

for i in range(5, 14169):
    values = []
    for j in range(1, 101):  # Columns 1 to 14168
        
        if j in columnsToAvoid:
            continue 
        if j == 1:
            value = int(sheet.cell(row=i, column=j).value)
        elif j == 2 or j == 3 or j == 100:
            value = sheet.cell(row=i, column=j).value
        else:
            value = get_cell_value(i, j)
        values.append(value)
    
    cursor.execute(sql, tuple(values))
    
    mydb.commit()

print("done")